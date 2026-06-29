#!/usr/bin/env python3
"""
Export latest yt2mp3 melody/fusion CSV+JSON artifacts to an Excel report, excluding raw_f0_hz from official comparisons.

Place:
    scripts/export_melody_excel_v4_2.py

Run:
    python scripts/export_melody_excel_v4_2.py

Default behavior:
    - auto-detect repo root
    - scan repo + actual runtime tmp outputs such as /tmp/yks/<job_id>
    - ignore pytest temp artifacts
    - choose the newest valid job group containing CSV + JSON
    - write Excel to repo/tests/output/fusion_report_YYYYMMDD_HHMMSS.xlsx

Optional:
    python scripts/export_melody_excel_v4_2.py --artifact-root /tmp/yks
    python scripts/export_melody_excel_v4_2.py --artifact-root /tmp/yks/<job_id>
    python scripts/export_melody_excel_v4_2.py --out tests/output/my_report.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import statistics
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

SCRIPT_VERSION = "export_melody_excel_v4_2_hybrid_compare_20260629"

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl\n"
        "Install:\n"
        "  python -m pip install openpyxl\n"
        "Suggested dev requirement:\n"
        "  openpyxl>=3.1\n"
    ) from exc


EXCLUDE_DIRS = {
    ".git", ".hg", ".svn",
    ".venv", "venv", "env",
    "node_modules", "__pycache__",
    ".pytest_cache", ".mypy_cache", ".ruff_cache",
    "dist", "build",
}

TIME_COLUMNS = (
    "time_sec", "time_seconds", "time", "seconds", "sec", "timestamp", "t",
    "start_sec", "frame_time", "frame_sec",
)

PITCH_HINTS = (
    "midi", "f0", "pitch", "frequency", "freq", "hz", "note_number",
)

IGNORE_PITCH_HINTS = (
    "confidence", "conf", "prob", "score", "voiced",
    "delta", "diff", "range", "error", "duration",
    "start", "end", "beat", "onset", "frame", "index", "id",
)

RAW_PITCH_COLUMN_NAMES = {
    "raw_f0_hz",
    "raw_f0",
    "raw_pitch_hz",
    "raw_frequency_hz",
    "raw_freq_hz",
}

RAW_PITCH_COLUMN_PATTERNS = (
    re.compile(r"(^|[_\-\.])raw[_\-\.]?f0($|[_\-\.])", re.I),
    re.compile(r"(^|[_\-\.])raw[_\-\.]?f0[_\-\.]?hz($|[_\-\.])", re.I),
    re.compile(r"(^|[_\-\.])raw[_\-\.]?pitch($|[_\-\.])", re.I),
)


def is_raw_pitch_column(header: str) -> bool:
    """True for diagnostic/raw pitch columns that should not enter official overlay/disagreement."""
    lower = header.lower().strip()
    normalized = re.sub(r"[^a-z0-9]+", "_", lower).strip("_")
    if lower in RAW_PITCH_COLUMN_NAMES or normalized in RAW_PITCH_COLUMN_NAMES:
        return True
    return any(pattern.search(lower) or pattern.search(normalized) for pattern in RAW_PITCH_COLUMN_PATTERNS)



CONF_HINTS = ("confidence", "conf", "prob", "score", "voicing", "voiced_prob")

CSV_NAME_BONUS = (
    "fusion", "comparison", "postprocess", "diagnostic", "pitch", "melody",
    "rmvpe", "crepe", "fcpe", "pesto",
)

MODEL_NAME_PATTERNS = {
    "rmvpe": re.compile(r"rmvpe", re.I),
    "torchcrepe": re.compile(r"torch[-_]?crepe|crepe", re.I),
    "fcpe": re.compile(r"fcpe", re.I),
    "pesto": re.compile(r"pesto", re.I),
    "fusion": re.compile(r"fusion|fused", re.I),
    "comparison": re.compile(r"comparison|compare", re.I),
}

YOUTUBE_METADATA_KEYS = {
    "id", "title", "fulltitle", "webpage_url", "original_url", "url",
    "uploader", "channel", "channel_id", "duration", "duration_string",
    "upload_date", "thumbnail", "extractor", "extractor_key",
}


OFFICIAL_MODEL_ORDER = ("fusion", "fusion_postprocessed", "rmvpe", "rmvpe_postprocessed", "hybrid_postprocessed", "torchcrepe", "fcpe", "pesto")

OFFICIAL_MODEL_KEYWORDS = {
    "fusion": ("fusion", "fused"),
    "fusion_postprocessed": ("fusion_postprocessed", "fusion-postprocessed", "fusion postprocessed"),
    "rmvpe": ("rmvpe",),
    "rmvpe_postprocessed": ("rmvpe_postprocessed", "rmvpe-postprocessed", "rmvpe postprocessed"),
    "hybrid_postprocessed": ("hybrid_postprocessed", "hybrid-postprocessed", "hybrid postprocessed"),
    "torchcrepe": ("torchcrepe", "torch_crepe", "torch-crepe"),
    "fcpe": ("fcpe",),
    "pesto": ("pesto",),
}

OFFICIAL_PITCH_COLUMN_PREFERENCE = (
    "f0_hz",
    "pitch_hz",
    "frequency_hz",
    "freq_hz",
    "hz",
    "midi",
    "note_number",
    "pitch",
    "f0",
)

V4_MODEL_DISAGREEMENT_TRACKS = ("fusion", "rmvpe", "torchcrepe", "fcpe", "pesto")
V4_FINAL_COMPARE_TRACKS = ("fusion_postprocessed", "rmvpe_postprocessed", "hybrid_postprocessed")
V4_POSTPROCESS_DELTA_PAIRS = (
    ("fusion", "fusion_postprocessed"),
    ("rmvpe", "rmvpe_postprocessed"),
)



@dataclass
class ArtifactGroup:
    group_root: Path
    analysis_dir: Path | None = None
    csv_files: list[Path] = field(default_factory=list)
    json_files: list[Path] = field(default_factory=list)
    score: float = 0.0
    newest_mtime: float = 0.0


@dataclass
class CsvSummary:
    path: Path
    row_count: int
    headers: list[str]
    time_column: str | None
    pitch_columns: list[str]
    confidence_columns: list[str]


@dataclass
class PitchTrack:
    label: str
    source_file: Path
    source_column: str
    values: list[tuple[float, float]]
    confidence_column: str | None = None


def log(message: str, level: str = "INFO") -> None:
    icon = {"INFO": "🟩", "WARN": "🟨", "ERROR": "🟥", "DEBUG": "🟦"}.get(level, "🟦")
    now = datetime.now().strftime("%H:%M:%S")
    print(f"{icon} [{now}][{level}] {message}")


def find_repo_root(start: Path | None = None) -> Path:
    cur = (start or Path.cwd()).resolve()
    for candidate in (cur, *cur.parents):
        if (candidate / ".git").exists() or (candidate / "pyproject.toml").exists() or (candidate / "README.md").exists():
            return candidate
    return cur


def safe_rel(path: Path, repo_root: Path) -> str:
    try:
        return str(path.resolve().relative_to(repo_root.resolve()))
    except Exception:
        return str(path)


def add_unique_path(paths: list[Path], value: str | Path | None) -> None:
    if not value:
        return
    p = Path(value).expanduser().resolve()
    if p.exists() and p not in paths:
        paths.append(p)


def build_search_roots(repo_root: Path, artifact_roots: list[Path]) -> list[Path]:
    roots: list[Path] = []
    add_unique_path(roots, repo_root)

    # CLI roots should be scanned early.
    for root in artifact_roots:
        add_unique_path(roots, root)

    # Env roots, if your app sets one later.
    for env_name in (
        "YT2MP3_ARTIFACT_ROOT",
        "YT2MP3_OUTPUT_DIR",
        "YT2MP3_JOB_DIR",
        "JOB_OUTPUT_DIR",
        "TMPDIR",
    ):
        add_unique_path(roots, os.getenv(env_name))

    # Actual runtime layout observed from your machine:
    # /tmp/yks/<uuid>/analysis/melody/fusion/...
    for common in (
        "/tmp/yks",
        "/tmp/yt2mp3",
        "/tmp/yt2mp3-fusion-debug",
        "/tmp/yt2mp3-jobs",
        "/tmp/yt2mp3_jobs",
        "/tmp/jobs",
    ):
        add_unique_path(roots, common)

    return roots


def is_ignored_path(path: Path) -> bool:
    text = str(path).replace("\\", "/").lower()
    if "/tmp/pytest-" in text or "/tmp/pytest-of-" in text:
        return True
    if "/node_modules/" in text or "/site-packages/" in text:
        return True
    if "package-lock" in text:
        return True
    return False


def iter_files(root: Path, suffixes: tuple[str, ...]) -> Iterable[Path]:
    if not root.exists():
        return
    for dirpath, dirnames, filenames in os.walk(root):
        current = Path(dirpath)
        if is_ignored_path(current):
            dirnames[:] = []
            continue
        dirnames[:] = [
            d for d in dirnames
            if d not in EXCLUDE_DIRS
            and not d.startswith(".")
            and not d.startswith("pytest-")
            and d != "pytest-of-startech"
        ]
        for name in filenames:
            if name.startswith("."):
                continue
            path = current / name
            if path.suffix.lower() in suffixes and not is_ignored_path(path):
                yield path


def choose_group_root(path: Path) -> tuple[Path, Path | None]:
    parts_lower = [p.lower() for p in path.parts]
    if "analysis" in parts_lower:
        idx = parts_lower.index("analysis")
        analysis_dir = Path(*path.parts[:idx + 1])
        return analysis_dir.parent, analysis_dir

    # Postprocess artifacts live at <artifact-root>/postprocess/...
    if "postprocess" in parts_lower:
        idx = parts_lower.index("postprocess")
        if idx > 0:
            return Path(*path.parts[:idx]), None

    # Fallback for /tmp/yt2mp3-fusion-debug/fusion.csv style.
    interesting = {"melody", "pitch", "fusion", "diagnostics", "debug", "inputs", "postprocess"}
    for idx, part in enumerate(parts_lower):
        if part in interesting and idx > 0:
            return Path(*path.parts[:idx]), None

    return path.parent, None


def discover_latest_group(search_roots: list[Path]) -> ArtifactGroup:
    groups: dict[Path, ArtifactGroup] = {}
    scanned = 0

    for root in search_roots:
        for path in iter_files(root, (".csv", ".json")):
            scanned += 1
            lower = str(path).replace("\\", "/").lower()
            if path.name.lower() in {"package.json", "tsconfig.json"}:
                continue

            # Keep likely yt2mp3 artifacts only.
            if not any(k in lower for k in (
                "analysis", "melody", "fusion", "pitch", "comparison", "postprocess", "rmvpe",
                "fcpe", "pesto", "crepe", "diagnostic", "metadata",
            )):
                continue

            group_root, analysis_dir = choose_group_root(path)
            group = groups.setdefault(group_root, ArtifactGroup(group_root=group_root, analysis_dir=analysis_dir))
            if analysis_dir and group.analysis_dir is None:
                group.analysis_dir = analysis_dir
            if path.suffix.lower() == ".csv":
                group.csv_files.append(path)
            elif path.suffix.lower() == ".json":
                group.json_files.append(path)
            try:
                group.newest_mtime = max(group.newest_mtime, path.stat().st_mtime)
            except OSError:
                pass

    candidates: list[ArtifactGroup] = []
    for group in groups.values():
        if not group.csv_files or not group.json_files:
            continue

        text = str(group.group_root).replace("\\", "/").lower()
        name_bonus = 0
        for p in [*group.csv_files, *group.json_files]:
            name_bonus += sum(1 for key in CSV_NAME_BONUS if key in p.name.lower())

        structure_bonus = 0
        if "/tmp/yks/" in text:
            structure_bonus += 200
        if (group.group_root / "analysis" / "melody" / "fusion" / "fusion.csv").exists():
            structure_bonus += 300
        if (group.group_root / "analysis" / "melody" / "fusion" / "diagnostics.json").exists():
            structure_bonus += 150
        if (group.group_root / "analysis" / "melody" / "fusion" / "inputs" / "rmvpe.csv").exists():
            structure_bonus += 50
        if (group.group_root / "analysis" / "stems" / "metadata.json").exists():
            structure_bonus += 50
        if (group.group_root / "postprocess" / "comparison_postprocessed.csv").exists():
            structure_bonus += 260
        if (group.group_root / "postprocess" / "postprocess_diagnostics.json").exists():
            structure_bonus += 120

        group.score = group.newest_mtime + name_bonus * 1000 + structure_bonus
        candidates.append(group)

    if not candidates:
        roots_text = "\n".join(f"  - {p}" for p in search_roots)
        raise SystemExit(
            "找不到同時包含 CSV 與 JSON 的 melody/fusion 分析資料。\n"
            "已搜尋 artifact search roots:\n"
            f"{roots_text}\n\n"
            "請確認是否存在：/tmp/yks/<job_id>/analysis/melody/fusion/fusion.csv"
        )

    candidates.sort(key=lambda g: (g.score, g.newest_mtime), reverse=True)
    return candidates[0]


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        return n if math.isfinite(n) else None
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "inf", "-inf"}:
        return None
    try:
        n = float(text)
        return n if math.isfinite(n) else None
    except ValueError:
        return None


def hz_to_midi(hz: float) -> float | None:
    if hz <= 0 or not math.isfinite(hz):
        return None
    return 69.0 + 12.0 * math.log2(hz / 440.0)


def looks_like_hz(column: str, values: list[float]) -> bool:
    lower = column.lower()
    if "midi" in lower or "note_number" in lower:
        return False
    if "hz" in lower or "f0" in lower or "freq" in lower or "frequency" in lower:
        return True
    if not values:
        return False
    try:
        med = statistics.median(values)
    except statistics.StatisticsError:
        return False
    return med > 125


def choose_time_column(headers: list[str]) -> str | None:
    lower_map = {h.lower().strip(): h for h in headers}
    for candidate in TIME_COLUMNS:
        if candidate in lower_map:
            return lower_map[candidate]
    for h in headers:
        lower = h.lower().strip()
        if "time" in lower and "timeout" not in lower:
            return h
    return None


def is_confidence_column(header: str) -> bool:
    lower = header.lower().strip()
    return any(h in lower for h in CONF_HINTS)


def is_pitch_column(header: str) -> bool:
    if is_raw_pitch_column(header):
        return False
    lower = header.lower().strip()
    tokens = {tok for tok in re.split(r"[^a-z0-9]+", lower) if tok}
    if tokens & {"id", "index", "frame"}:
        return False
    for skip in IGNORE_PITCH_HINTS:
        if skip in {"id", "index", "frame"}:
            continue
        if skip in lower:
            return False
    return any(h in lower for h in PITCH_HINTS)


def read_csv_rows(path: Path, limit: int | None = None) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        rows = []
        for i, row in enumerate(reader):
            if limit is not None and i >= limit:
                break
            rows.append(row)
    return headers, rows


def summarize_csv(path: Path) -> CsvSummary:
    try:
        headers, rows = read_csv_rows(path, limit=5000)
    except Exception as exc:
        log(f"CSV read failed, skipped: {path} ({exc})", "WARN")
        return CsvSummary(path=path, row_count=0, headers=[], time_column=None, pitch_columns=[], confidence_columns=[])

    time_col = choose_time_column(headers)
    numeric_headers: list[str] = []
    threshold = max(3, min(10, len(rows) // 10 if rows else 3))
    for h in headers:
        count = 0
        for row in rows[:200]:
            if safe_float(row.get(h)) is not None:
                count += 1
        if count >= threshold:
            numeric_headers.append(h)

    pitch_cols = [h for h in numeric_headers if is_pitch_column(h)]
    conf_cols = [h for h in numeric_headers if is_confidence_column(h)]

    # Fallback for simple model CSVs where columns may be named e.g. value.
    if not pitch_cols and time_col:
        filename = path.name.lower()
        if any(k in filename for k in ("rmvpe", "crepe", "fcpe", "pesto", "pitch", "fusion")):
            pitch_cols = [
                h for h in numeric_headers
                if h != time_col and not is_confidence_column(h)
                and not is_raw_pitch_column(h)
                and not any(skip in h.lower() for skip in IGNORE_PITCH_HINTS)
            ][:3]

    return CsvSummary(path=path, row_count=len(rows), headers=headers, time_column=time_col, pitch_columns=pitch_cols, confidence_columns=conf_cols)


def priority_for_csv(path: Path) -> int:
    name = path.name.lower()
    full = str(path).replace("\\", "/").lower()
    if "/inputs/" in full and any(k in name for k in ("rmvpe", "torchcrepe", "fcpe", "pesto")):
        return 0
    if "fusion.csv" == name:
        return 1
    if "comparison" in name or "compare" in name:
        return 2
    if "fusion" in name or "fused" in name:
        return 3
    return 9


def normalize_label(path: Path, column: str) -> str:
    source = f"{path.parent.name}_{path.stem}_{column}".lower()
    model = None
    for name, pattern in MODEL_NAME_PATTERNS.items():
        if pattern.search(source):
            model = name
            break
    if model:
        col = column.lower()
        if col in {"midi", "f0", "f0_hz", "pitch", "pitch_hz", "frequency", "hz"}:
            return model
        return f"{model}_{column}"
    return f"{path.stem}_{column}"


def guess_confidence_column(pitch_col: str, conf_cols: list[str]) -> str | None:
    if not conf_cols:
        return None
    pitch_lower = pitch_col.lower()
    for conf in conf_cols:
        conf_lower = conf.lower()
        for prefix in ("rmvpe", "torchcrepe", "crepe", "fcpe", "pesto", "fusion", "fused"):
            if prefix in pitch_lower and prefix in conf_lower:
                return conf
    return conf_cols[0]


def extract_tracks(csv_files: list[Path]) -> tuple[list[PitchTrack], list[CsvSummary]]:
    tracks: list[PitchTrack] = []
    summaries: list[CsvSummary] = []
    seen: dict[str, int] = {}

    for path in sorted(csv_files, key=lambda p: (priority_for_csv(p), str(p))):
        summary = summarize_csv(path)
        summaries.append(summary)
        if not summary.time_column or not summary.pitch_columns:
            continue

        try:
            _, rows = read_csv_rows(path)
        except Exception:
            continue

        for pitch_col in summary.pitch_columns:
            if is_raw_pitch_column(pitch_col):
                continue
            raw: list[tuple[float, float]] = []
            raw_values: list[float] = []
            for row in rows:
                t = safe_float(row.get(summary.time_column))
                y = safe_float(row.get(pitch_col))
                if t is None or y is None or y <= 0:
                    continue
                raw.append((t, y))
                raw_values.append(y)

            if not raw:
                continue

            convert_hz = looks_like_hz(pitch_col, raw_values)
            points: list[tuple[float, float]] = []
            for t, y in raw:
                midi = hz_to_midi(y) if convert_hz else y
                if midi is None:
                    continue
                if 20 <= midi <= 110:
                    points.append((float(t), float(midi)))

            if not points:
                continue

            label = normalize_label(path, pitch_col)
            if label in seen:
                seen[label] += 1
                label = f"{label}_{seen[label]}"
            else:
                seen[label] = 1

            tracks.append(
                PitchTrack(
                    label=label,
                    source_file=path,
                    source_column=pitch_col,
                    values=points,
                    confidence_column=guess_confidence_column(pitch_col, summary.confidence_columns),
                )
            )

    return tracks, summaries


def find_comparison_csv(csv_files: list[Path]) -> Path | None:
    """Prefer postprocessed comparison CSV, then aligned comparison.csv."""
    candidates = []
    for path in csv_files:
        name = path.name.lower()
        full = str(path).replace("\\", "/").lower()
        if name == "comparison_postprocessed.csv":
            candidates.append((0 if "/postprocess/" in full else 1, path))
        elif name == "comparison.csv":
            candidates.append((10 if "/fusion/" in full else 11, path))
        elif "comparison" in name or "compare" in name:
            candidates.append((20, path))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], str(item[1])))
    return candidates[0][1]


def model_matches_column(model: str, column: str) -> bool:
    lower = column.lower().strip()
    return any(keyword in lower for keyword in OFFICIAL_MODEL_KEYWORDS.get(model, (model,)))


def pitch_column_preference_score(column: str) -> int:
    lower = column.lower().strip()
    for idx, hint in enumerate(OFFICIAL_PITCH_COLUMN_PREFERENCE):
        if hint in lower:
            return idx
    return len(OFFICIAL_PITCH_COLUMN_PREFERENCE) + 10


def choose_official_pitch_column(model: str, headers: list[str], time_col: str | None) -> str | None:
    candidates = []
    for h in headers:
        if h == time_col:
            continue
        if is_raw_pitch_column(h):
            continue
        if is_confidence_column(h):
            continue
        if not is_pitch_column(h):
            continue
        if not model_matches_column(model, h):
            continue
        lower = h.lower().strip()

        # Avoid columns that are clearly derived errors/ranges, even if they include f0-like names.
        if any(skip in lower for skip in ("abs_cents", "cents_error", "error", "delta", "diff", "range", "spread")):
            continue

        # Prefer the canonical model_f0_hz / fusion_f0_hz style.
        exact_bonus = 0
        for keyword in OFFICIAL_MODEL_KEYWORDS.get(model, (model,)):
            normalized = re.sub(r"[^a-z0-9]+", "_", lower).strip("_")
            k_norm = re.sub(r"[^a-z0-9]+", "_", keyword).strip("_")
            if normalized in {
                f"{k_norm}_f0_hz",
                f"{k_norm}_pitch_hz",
                f"{k_norm}_midi",
                f"{k_norm}_f0",
            }:
                exact_bonus -= 10
                break

        candidates.append((exact_bonus + pitch_column_preference_score(h), len(h), h))

    if not candidates:
        return None
    candidates.sort()
    return candidates[0][2]


def make_pitch_track_from_column(label: str, path: Path, time_col: str, pitch_col: str, rows: list[dict[str, str]], conf_cols: list[str]) -> PitchTrack | None:
    raw: list[tuple[float, float]] = []
    raw_values: list[float] = []
    for row in rows:
        t = safe_float(row.get(time_col))
        y = safe_float(row.get(pitch_col))
        if t is None or y is None or y <= 0:
            continue
        raw.append((t, y))
        raw_values.append(y)

    if not raw:
        return None

    convert_hz = looks_like_hz(pitch_col, raw_values)
    points: list[tuple[float, float]] = []
    for t, y in raw:
        midi = hz_to_midi(y) if convert_hz else y
        if midi is None:
            continue
        if 20 <= midi <= 110:
            points.append((float(t), float(midi)))

    if not points:
        return None

    return PitchTrack(
        label=label,
        source_file=path,
        source_column=pitch_col,
        values=points,
        confidence_column=guess_confidence_column(pitch_col, conf_cols),
    )


def extract_official_tracks(csv_files: list[Path]) -> tuple[list[PitchTrack], str]:
    """
    Official overlay/disagreement source.

    First choice:
      comparison.csv, because all model/fusion columns are aligned on the same time grid.

    Fallback:
      deduplicated tracks from all CSVs, keeping only one line per model.
    """
    comparison = find_comparison_csv(csv_files)
    if comparison is not None:
        summary = summarize_csv(comparison)
        if summary.time_column:
            try:
                headers, rows = read_csv_rows(comparison)
            except Exception:
                headers, rows = [], []

            tracks: list[PitchTrack] = []
            for model in OFFICIAL_MODEL_ORDER:
                col = choose_official_pitch_column(model, headers, summary.time_column)
                if col is None:
                    continue
                track = make_pitch_track_from_column(
                    label=model,
                    path=comparison,
                    time_col=summary.time_column,
                    pitch_col=col,
                    rows=rows,
                    conf_cols=summary.confidence_columns,
                )
                if track is not None:
                    tracks.append(track)

            if tracks:
                return tracks, f"{comparison.name}: {comparison}"

    # Fallback: use previous broad extraction, but deduplicate to official model names.
    all_tracks, _ = extract_tracks(csv_files)
    selected: dict[str, PitchTrack] = {}
    for model in OFFICIAL_MODEL_ORDER:
        for track in all_tracks:
            source = f"{track.label} {track.source_file.name} {track.source_column}".lower()
            if any(keyword in source for keyword in OFFICIAL_MODEL_KEYWORDS.get(model, (model,))):
                if model not in selected or priority_for_csv(track.source_file) < priority_for_csv(selected[model].source_file):
                    selected[model] = PitchTrack(
                        label=model,
                        source_file=track.source_file,
                        source_column=track.source_column,
                        values=track.values,
                        confidence_column=track.confidence_column,
                    )

    tracks = [selected[m] for m in OFFICIAL_MODEL_ORDER if m in selected]
    return tracks, "deduplicated CSV fallback"


def extract_row_aligned_official_rows(csv_files: list[Path]) -> tuple[list[str], list[dict[str, Any]], str]:
    """
    Load the official comparison source as row-aligned MIDI rows.

    Important:
      Pitch Overlay is downsampled for chart readability.
      Diagnostic sheets must not use Pitch Overlay, because each track may be
      downsampled at different time points. This function reads
      comparison_postprocessed.csv / comparison.csv directly, preserving row
      alignment.
    """
    comparison = find_comparison_csv(csv_files)
    if comparison is None:
        return [], [], "row-aligned source not found"

    summary = summarize_csv(comparison)
    if not summary.time_column:
        return [], [], f"{comparison.name}: missing time column"

    try:
        headers, raw_rows = read_csv_rows(comparison)
    except Exception as exc:
        return [], [], f"{comparison.name}: read failed ({exc})"

    model_columns: dict[str, str] = {}
    convert_hz_by_model: dict[str, bool] = {}

    for model in OFFICIAL_MODEL_ORDER:
        col = choose_official_pitch_column(model, headers, summary.time_column)
        if col is None:
            continue
        model_columns[model] = col

        sample_values: list[float] = []
        for row in raw_rows[:5000]:
            y = safe_float(row.get(col))
            if y is not None and y > 0:
                sample_values.append(y)
        convert_hz_by_model[model] = looks_like_hz(col, sample_values)

    out_headers = ["time_sec"] + [model for model in OFFICIAL_MODEL_ORDER if model in model_columns]
    out_rows: list[dict[str, Any]] = []

    for idx, row in enumerate(raw_rows):
        t = safe_float(row.get(summary.time_column))
        if t is None:
            continue

        out: dict[str, Any] = {"time_sec": round(float(t), 4)}
        for model in out_headers[1:]:
            col = model_columns[model]
            y = safe_float(row.get(col))
            if y is None or y <= 0:
                out[model] = ""
                continue

            midi = hz_to_midi(y) if convert_hz_by_model.get(model, False) else y
            if midi is None or midi < 20 or midi > 110:
                out[model] = ""
            else:
                out[model] = round(float(midi), 6)

        # Keep the row if any pitch track has a value.
        if any(safe_float(out.get(model)) is not None for model in out_headers[1:]):
            out_rows.append(out)

    source = f"{comparison.name}: {comparison}"
    return out_headers, out_rows, source


def build_octave_candidate_rows(headers: list[str], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    pitch_cols = [h for h in headers if h != "time_sec"]
    candidates: list[dict[str, Any]] = []
    octave_targets = (1200, 2400, 3600)

    for row in rows:
        values = []
        for col in pitch_cols:
            y = safe_float(row.get(col))
            if y is not None:
                values.append((col, y))
        if len(values) < 2:
            continue

        all_models = ", ".join(f"{name}={midi:.2f}" for name, midi in values)
        for i, (a_name, a_midi) in enumerate(values):
            for b_name, b_midi in values[i + 1:]:
                diff_cents = abs(a_midi - b_midi) * 100.0
                target = min(octave_targets, key=lambda t: abs(diff_cents - t))
                error = abs(diff_cents - target)

                # A permissive window is useful for real vocal material and model jitter.
                if error <= 125:
                    candidates.append({
                        "time_sec": row.get("time_sec"),
                        "model_a": a_name,
                        "midi_a": round(a_midi, 4),
                        "model_b": b_name,
                        "midi_b": round(b_midi, 4),
                        "diff_cents": round(diff_cents, 2),
                        "nearest_octave_cents": target,
                        "octave_error_cents": round(error, 2),
                        "all_models": all_models,
                    })

    candidates.sort(key=lambda r: (float(r["octave_error_cents"]), -float(r["diff_cents"])))
    return candidates[:500]


def downsample(points: list[tuple[float, float]], max_points: int) -> list[tuple[float, float]]:
    if len(points) <= max_points:
        return points
    step = max(1, math.ceil(len(points) / max_points))
    return points[::step]


def build_overlay_rows(tracks: list[PitchTrack], max_points: int) -> tuple[list[str], list[dict[str, Any]]]:
    overlay: dict[float, dict[str, Any]] = {}
    for track in tracks:
        for t, midi in downsample(track.values, max_points):
            key = round(t, 3)
            row = overlay.setdefault(key, {"time_sec": key})
            row[track.label] = round(midi, 4)

    rows = [overlay[k] for k in sorted(overlay)]
    if len(rows) > max_points:
        step = max(1, math.ceil(len(rows) / max_points))
        rows = rows[::step]
    headers = ["time_sec"] + [t.label for t in tracks]
    return headers, rows


def build_disagreement_rows(headers: list[str], rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    pitch_cols = [h for h in headers if h != "time_sec"]
    out: list[dict[str, Any]] = []
    for row in rows:
        values = []
        for col in pitch_cols:
            y = safe_float(row.get(col))
            if y is not None:
                values.append((col, y))
        if len(values) < 2:
            continue
        ys = [y for _, y in values]
        spread_cents = (max(ys) - min(ys)) * 100
        if spread_cents < 50:
            continue
        # crude octave-pair detector
        octave_like = False
        for i, (_, a) in enumerate(values):
            for _, b in values[i + 1:]:
                if abs(abs(a - b) - 12) <= 0.75:
                    octave_like = True
        out.append({
            "time_sec": row.get("time_sec"),
            "support_count": len(values),
            "min_midi": round(min(ys), 4),
            "max_midi": round(max(ys), 4),
            "spread_cents": round(spread_cents, 2),
            "suspected_octave_pair": "Y" if octave_like else "",
            "models": ", ".join(name for name, _ in values),
        })
    out.sort(key=lambda r: float(r["spread_cents"]), reverse=True)
    return out[:500]


def load_json(path: Path) -> Any | None:
    try:
        with path.open("r", encoding="utf-8-sig") as f:
            return json.load(f)
    except Exception:
        return None


def flatten_dict(obj: Any, prefix: str = "", limit: int = 300) -> dict[str, Any]:
    out: dict[str, Any] = {}

    def rec(value: Any, key: str) -> None:
        if len(out) >= limit:
            return
        if isinstance(value, dict):
            for k, v in value.items():
                child = f"{key}.{k}" if key else str(k)
                rec(v, child)
        elif isinstance(value, list):
            if value and len(value) <= 5:
                for i, v in enumerate(value):
                    rec(v, f"{key}[{i}]")
            else:
                out[key] = f"[list len={len(value)}]"
        else:
            out[key] = value

    rec(obj, prefix)
    return out


def metadata_score(flat: dict[str, Any]) -> int:
    score = 0
    for key in flat:
        short = key.split(".")[-1]
        if short in YOUTUBE_METADATA_KEYS:
            score += 10
    for key in ("title", "fulltitle"):
        if any(k.endswith(key) or k == key for k in flat):
            score += 30
    if any("webpage_url" in k or "youtube" in str(v).lower() for k, v in flat.items()):
        score += 30
    return score


def pick_metadata(json_files: list[Path]) -> tuple[Path | None, dict[str, Any]]:
    best_path: Path | None = None
    best_flat: dict[str, Any] = {}
    best_score = -1
    for path in json_files:
        data = load_json(path)
        if data is None:
            continue
        flat = flatten_dict(data)
        score = metadata_score(flat)
        # filename hints
        name = path.name.lower()
        if "metadata" in name or "info" in name:
            score += 20
        if score > best_score:
            best_score = score
            best_path = path
            best_flat = flat
    return best_path, best_flat


def meta_get(metadata: dict[str, Any], names: tuple[str, ...]) -> Any:
    # Prefer exact short keys first.
    for name in names:
        if name in metadata and metadata[name] not in (None, ""):
            return metadata[name]
    for key, value in metadata.items():
        short = key.split(".")[-1]
        if short in names and value not in (None, ""):
            return value
    return ""


def format_upload_date(value: Any) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"\d{8}", text):
        return f"{text[:4]}-{text[4:6]}-{text[6:]}"
    return text


def set_basic_styles(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(60, len(str(cell.value))))
        ws.column_dimensions[letter].width = min(45, max_len + 2)
    ws.freeze_panes = "A2"


def append_kv(ws, key: str, value: Any) -> None:
    ws.append([key, "" if value is None else value])


def find_flat_metric(flat: dict[str, Any], model: str, aliases: tuple[str, ...]) -> tuple[Any, str]:
    best: tuple[int, str, Any] | None = None
    model_l = model.lower()
    for key, value in flat.items():
        lower = key.lower()
        if model_l not in lower:
            continue
        if not any(alias in lower for alias in aliases):
            continue

        # Prefer keys that end close to the metric name and are shorter.
        score = len(lower)
        for alias in aliases:
            if lower.endswith(alias):
                score -= 100
        candidate = (score, key, value)
        if best is None or candidate < best:
            best = candidate

    if best is None:
        return "", ""
    return best[2], best[1]


def build_model_summary_rows(json_files: list[Path]) -> list[dict[str, Any]]:
    combined: dict[str, Any] = {}
    source_by_key: dict[str, str] = {}

    for path in json_files:
        data = load_json(path)
        if data is None:
            continue
        flat = flatten_dict(data, limit=2000)
        for key, value in flat.items():
            full_key = f"{path.name}:{key}"
            combined[full_key] = value
            source_by_key[full_key] = str(path)

    rows: list[dict[str, Any]] = []
    metric_aliases = {
        "status": ("status",),
        "voiced_ratio": ("voiced_ratio", "voice_ratio"),
        "confidence_kind": ("confidence_kind", "confidence_type", "conf_kind"),
        "missing_confidence_rows": ("missing_confidence_rows", "missing_conf"),
        "weight": ("weight", "model_weight"),
        "common_voiced": ("common_voiced", "common_voiced_count"),
        "agree_50c_ratio": ("agree_50c_ratio", "agree_50_c_ratio", "agree_within_50"),
        "mean_abs_cents": ("mean_abs_cents", "mean_abs_cent", "mean_absolute_cents"),
    }

    for model in ("rmvpe", "torchcrepe", "fcpe", "pesto"):
        row: dict[str, Any] = {"model": model}
        source_keys = []
        for metric, aliases in metric_aliases.items():
            value, key = find_flat_metric(combined, model, aliases)
            row[metric] = value
            if key:
                source_keys.append(key)
        row["source_keys"] = "\n".join(source_keys[:6])
        rows.append(row)

    return rows


def write_summary(wb: Workbook, group: ArtifactGroup, repo_root: Path, metadata_path: Path | None, metadata: dict[str, Any], tracks: list[PitchTrack], disagreement_rows: list[dict[str, Any]], official_source: str) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    append_kv(ws, "Script version", SCRIPT_VERSION)
    append_kv(ws, "Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    append_kv(ws, "Job root", str(group.group_root))
    append_kv(ws, "Analysis dir", str(group.analysis_dir or ""))
    append_kv(ws, "Metadata source", str(metadata_path or ""))

    append_kv(ws, "Title", meta_get(metadata, ("title", "fulltitle")))
    append_kv(ws, "Channel / uploader", meta_get(metadata, ("channel", "uploader")))
    append_kv(ws, "Video ID", meta_get(metadata, ("id", "display_id")))
    append_kv(ws, "Duration", meta_get(metadata, ("duration_string", "duration")))
    append_kv(ws, "Upload date", format_upload_date(meta_get(metadata, ("upload_date",))))
    append_kv(ws, "YouTube URL", meta_get(metadata, ("webpage_url", "original_url", "url")))
    append_kv(ws, "Thumbnail", meta_get(metadata, ("thumbnail",)))
    append_kv(ws, "Extractor", meta_get(metadata, ("extractor", "extractor_key")))

    append_kv(ws, "CSV files found", len(group.csv_files))
    append_kv(ws, "JSON files found", len(group.json_files))
    append_kv(ws, "Pitch tracks exported", len(tracks))
    append_kv(ws, "Official comparison source", official_source)
    append_kv(ws, "Official tracks", ", ".join(track.label for track in tracks))
    append_kv(ws, "Excluded from official comparison", "raw_f0_hz, duplicated inputs/*.csv, duplicated fusion midi/f0 tracks")
    append_kv(ws, "Postprocess A/B lines", "fusion_postprocessed, rmvpe_postprocessed, hybrid_postprocessed when comparison_postprocessed.csv exists")
    append_kv(ws, "V4.2 split diagnostics", "Row-aligned CSV source, not downsampled Pitch Overlay")
    append_kv(ws, "V4 split diagnostics", "Model Disagreement / Postprocess Delta / Final Compare / Fusion Fill RMVPE Gap / Review Segments")
    append_kv(ws, "High-disagreement rows", len(disagreement_rows))
    set_basic_styles(ws)


def write_files_sheet(wb: Workbook, group: ArtifactGroup, repo_root: Path) -> None:
    ws = wb.create_sheet("Files")
    ws.append(["Type", "Modified", "Path"])
    files = [(p.suffix.lower().lstrip("."), p) for p in [*group.csv_files, *group.json_files]]
    files.sort(key=lambda x: x[1].stat().st_mtime if x[1].exists() else 0, reverse=True)
    for typ, path in files:
        try:
            mtime = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        except OSError:
            mtime = ""
        ws.append([typ, mtime, str(path)])
    set_basic_styles(ws)


def write_csv_index(wb: Workbook, summaries: list[CsvSummary]) -> None:
    ws = wb.create_sheet("CSV Index")
    ws.append(["File", "Rows", "Time column", "Pitch columns", "Confidence columns", "Headers"])
    for s in summaries:
        ws.append([
            str(s.path),
            s.row_count,
            s.time_column or "",
            ", ".join(s.pitch_columns),
            ", ".join(s.confidence_columns),
            ", ".join(s.headers),
        ])
    set_basic_styles(ws)


def write_overlay_sheet(wb: Workbook, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Pitch Overlay")
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    set_basic_styles(ws)

    if ws.max_row >= 3 and len(headers) >= 2:
        chart = LineChart()
        chart.title = "Pitch overlay (MIDI)"
        chart.y_axis.title = "MIDI"
        chart.x_axis.title = "Time sec"
        chart.height = 14
        chart.width = 28
        data = Reference(ws, min_col=2, max_col=min(ws.max_column, 10), min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J2")


def write_disagreement_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Disagreement")
    headers = ["time_sec", "support_count", "min_midi", "max_midi", "spread_cents", "suspected_octave_pair", "models"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    set_basic_styles(ws)

    if ws.max_row >= 3:
        chart = BarChart()
        chart.title = "Top pitch disagreements"
        chart.y_axis.title = "Spread cents"
        chart.x_axis.title = "Time sec"
        chart.height = 12
        chart.width = 24
        # Use top 30 rows only for chart readability.
        max_row = min(ws.max_row, 31)
        data = Reference(ws, min_col=5, min_row=1, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "I2")


def write_octave_candidates_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Octave Candidates")
    headers = [
        "time_sec",
        "model_a",
        "midi_a",
        "model_b",
        "midi_b",
        "diff_cents",
        "nearest_octave_cents",
        "octave_error_cents",
        "all_models",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    set_basic_styles(ws)

    if ws.max_row >= 3:
        chart = BarChart()
        chart.title = "Octave-like model disagreements"
        chart.y_axis.title = "Diff cents"
        chart.x_axis.title = "Time sec"
        chart.height = 12
        chart.width = 24
        max_row = min(ws.max_row, 31)
        data = Reference(ws, min_col=6, min_row=1, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "K2")


def write_model_summary_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Model Summary")
    headers = [
        "model",
        "status",
        "voiced_ratio",
        "confidence_kind",
        "missing_confidence_rows",
        "weight",
        "common_voiced",
        "agree_50c_ratio",
        "mean_abs_cents",
        "source_keys",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    set_basic_styles(ws)


def find_json_by_name(json_files: list[Path], filename: str) -> Path | None:
    matches = [path for path in json_files if path.name.lower() == filename.lower()]
    if not matches:
        return None
    matches.sort(key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)
    return matches[0]


def write_postprocess_summary_sheet(wb: Workbook, json_files: list[Path]) -> None:
    ws = wb.create_sheet("Postprocess Summary")
    ws.append(["Section", "Key", "Value"])

    post_diag_path = find_json_by_name(json_files, "postprocess_diagnostics.json")
    cleaned_path = find_json_by_name(json_files, "cleaned_melody.json")

    if post_diag_path is None and cleaned_path is None:
        ws.append(["status", "postprocess_artifacts", "not found"])
        set_basic_styles(ws)
        return

    if post_diag_path is not None:
        data = load_json(post_diag_path) or {}
        ws.append(["source", "postprocess_diagnostics", str(post_diag_path)])
        for key, value in (data.get("parameters") or {}).items():
            ws.append(["parameters", key, value])
        for target, target_data in (data.get("targets") or {}).items():
            stats = (target_data or {}).get("stats") or {}
            for key, value in stats.items():
                ws.append([f"{target}.stats", key, value])
            ws.append([f"{target}.details", "octave_corrections_preview", len((target_data or {}).get("octave_corrections") or [])])
            ws.append([f"{target}.details", "removed_short_islands_preview", len((target_data or {}).get("removed_short_islands") or [])])
        ws.append(["summary", "octave_candidates_count", data.get("octave_candidates_count", "")])
        ws.append(["summary", "estimated_hop_sec", data.get("estimated_hop_sec", "")])

    if cleaned_path is not None:
        data = load_json(cleaned_path) or {}
        ws.append(["source", "cleaned_melody", str(cleaned_path)])
        for track_name, track_data in (data.get("tracks") or {}).items():
            stats = (track_data or {}).get("stats") or {}
            ws.append([track_name, "mode", (track_data or {}).get("mode", "")])
            for key, value in stats.items():
                ws.append([f"{track_name}.stats", key, value])

    set_basic_styles(ws)


def filter_overlay_for_tracks(
    headers: list[str],
    rows: list[dict[str, Any]],
    keep_tracks: tuple[str, ...],
) -> tuple[list[str], list[dict[str, Any]]]:
    keep = ["time_sec"] + [track for track in keep_tracks if track in headers]
    filtered_rows = []
    for row in rows:
        new_row = {key: row.get(key, "") for key in keep}
        # Keep rows that have at least one pitch value besides time.
        if any(safe_float(new_row.get(key)) is not None for key in keep if key != "time_sec"):
            filtered_rows.append(new_row)
    return keep, filtered_rows


def build_postprocess_delta_rows(
    headers: list[str],
    rows: list[dict[str, Any]],
    threshold_cents: float = 50.0,
) -> list[dict[str, Any]]:
    deltas: list[dict[str, Any]] = []
    for row in rows:
        t = row.get("time_sec")
        for original_col, post_col in V4_POSTPROCESS_DELTA_PAIRS:
            if original_col not in headers or post_col not in headers:
                continue
            original = safe_float(row.get(original_col))
            post = safe_float(row.get(post_col))
            if original is None and post is None:
                continue
            if original is None or post is None:
                deltas.append({
                    "time_sec": t,
                    "target": original_col,
                    "original_midi": "" if original is None else round(original, 4),
                    "postprocessed_midi": "" if post is None else round(post, 4),
                    "delta_cents": "",
                    "change_type": "voiced_removed" if original is not None else "voiced_added",
                    "octave_like": "",
                })
                continue
            delta = abs(post - original) * 100.0
            if delta < threshold_cents:
                continue
            octave_like = any(abs(delta - target) <= 125 for target in (1200, 2400, 3600))
            deltas.append({
                "time_sec": t,
                "target": original_col,
                "original_midi": round(original, 4),
                "postprocessed_midi": round(post, 4),
                "delta_cents": round(delta, 2),
                "change_type": "octave_shift" if octave_like else "pitch_adjust",
                "octave_like": "Y" if octave_like else "",
            })
    deltas.sort(key=lambda r: float(r["delta_cents"] or 0), reverse=True)
    return deltas[:1000]


def build_final_compare_rows(
    headers: list[str],
    rows: list[dict[str, Any]],
    threshold_cents: float = 50.0,
) -> list[dict[str, Any]]:
    candidate_tracks = [track for track in V4_FINAL_COMPARE_TRACKS if track in headers]
    if len(candidate_tracks) < 2:
        return []

    out: list[dict[str, Any]] = []
    for row in rows:
        for i, a_name in enumerate(candidate_tracks):
            a = safe_float(row.get(a_name))
            for b_name in candidate_tracks[i + 1:]:
                b = safe_float(row.get(b_name))
                if a is None and b is None:
                    continue
                if a is None or b is None:
                    out.append({
                        "time_sec": row.get("time_sec"),
                        "candidate_a": a_name,
                        "candidate_b": b_name,
                        "midi_a": "" if a is None else round(a, 4),
                        "midi_b": "" if b is None else round(b, 4),
                        "diff_cents": "",
                        "relationship": f"{a_name}_only" if a is not None else f"{b_name}_only",
                        "octave_like": "",
                    })
                    continue

                diff = abs(a - b) * 100.0
                if diff < threshold_cents:
                    continue
                octave_like = any(abs(diff - target) <= 125 for target in (1200, 2400, 3600))
                out.append({
                    "time_sec": row.get("time_sec"),
                    "candidate_a": a_name,
                    "candidate_b": b_name,
                    "midi_a": round(a, 4),
                    "midi_b": round(b, 4),
                    "diff_cents": round(diff, 2),
                    "relationship": "octave_diff" if octave_like else "pitch_diff",
                    "octave_like": "Y" if octave_like else "",
                })

    out.sort(key=lambda r: float(r["diff_cents"] or 99999), reverse=True)
    return out[:1500]

def support_for_candidate(row: dict[str, Any], candidate: float, support_cents: float = 60.0) -> tuple[int, list[str]]:
    supporters = []
    for model in ("torchcrepe", "fcpe", "pesto"):
        y = safe_float(row.get(model))
        if y is not None and abs(y - candidate) * 100.0 <= support_cents:
            supporters.append(model)
    return len(supporters), supporters


def build_fusion_fill_gap_rows(
    headers: list[str],
    rows: list[dict[str, Any]],
    support_cents: float = 60.0,
) -> list[dict[str, Any]]:
    required = {"rmvpe", "fusion", "fusion_postprocessed"}
    if not required.issubset(set(headers)):
        return []

    gap_rows: list[dict[str, Any]] = []
    for row in rows:
        rmvpe = safe_float(row.get("rmvpe"))
        fusion = safe_float(row.get("fusion"))
        fusion_post = safe_float(row.get("fusion_postprocessed"))
        hybrid_post = safe_float(row.get("hybrid_postprocessed")) if "hybrid_postprocessed" in headers else None
        if rmvpe is not None:
            continue
        if fusion is None and fusion_post is None:
            continue

        candidate = fusion_post if fusion_post is not None else fusion
        if candidate is None:
            continue

        support_count, supporters = support_for_candidate(row, candidate, support_cents=support_cents)
        hybrid_decision = ""
        if "hybrid_postprocessed" in headers:
            hybrid_decision = "accepted" if hybrid_post is not None else "rejected"

        gap_rows.append({
            "time_sec": row.get("time_sec"),
            "fusion_midi": "" if fusion is None else round(fusion, 4),
            "fusion_postprocessed_midi": "" if fusion_post is None else round(fusion_post, 4),
            "hybrid_postprocessed_midi": "" if hybrid_post is None else round(hybrid_post, 4),
            "hybrid_decision": hybrid_decision,
            "support_count": support_count,
            "supporters": ", ".join(supporters),
            "torchcrepe_midi": row.get("torchcrepe", ""),
            "fcpe_midi": row.get("fcpe", ""),
            "pesto_midi": row.get("pesto", ""),
            "risk": "weak_support" if support_count == 0 else ("medium_support" if support_count == 1 else "supported"),
        })

    risk_rank = {"weak_support": 0, "medium_support": 1, "supported": 2}
    gap_rows.sort(key=lambda r: (risk_rank.get(str(r["risk"]), 9), str(r.get("hybrid_decision") or ""), float(r["time_sec"] or 0)))
    return gap_rows[:1000]

def build_review_segment_rows(
    model_disagreement_rows: list[dict[str, Any]],
    postprocess_delta_rows: list[dict[str, Any]],
    final_compare_rows: list[dict[str, Any]],
    fusion_fill_gap_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    review: list[dict[str, Any]] = []

    for row in fusion_fill_gap_rows[:200]:
        priority = 1 if row.get("risk") == "weak_support" else 2
        review.append({
            "priority": priority,
            "time_sec": row.get("time_sec"),
            "reason": "fusion fills RMVPE gap with weak/no support" if priority == 1 else "fusion fills RMVPE gap",
            "detail": f"risk={row.get('risk')}; decision={row.get('hybrid_decision')}; supporters={row.get('supporters')}; fusion_post={row.get('fusion_postprocessed_midi')}; hybrid={row.get('hybrid_postprocessed_midi')}",
        })

    for row in final_compare_rows[:200]:
        diff = safe_float(row.get("diff_cents"))
        priority = 1 if diff is not None and diff >= 900 else 2
        review.append({
            "priority": priority,
            "time_sec": row.get("time_sec"),
            "reason": "final RMVPE vs fusion differs by octave" if row.get("octave_like") == "Y" else "final RMVPE vs fusion differs",
            "detail": f"diff_cents={row.get('diff_cents')}; relationship={row.get('relationship')}",
        })

    for row in postprocess_delta_rows[:200]:
        diff = safe_float(row.get("delta_cents"))
        priority = 2 if row.get("octave_like") == "Y" or (diff is not None and diff >= 900) else 3
        review.append({
            "priority": priority,
            "time_sec": row.get("time_sec"),
            "reason": f"postprocess changed {row.get('target')}",
            "detail": f"change_type={row.get('change_type')}; delta_cents={row.get('delta_cents')}",
        })

    for row in model_disagreement_rows[:200]:
        spread = safe_float(row.get("spread_cents"))
        priority = 2 if spread is not None and spread >= 900 else 3
        review.append({
            "priority": priority,
            "time_sec": row.get("time_sec"),
            "reason": "raw model disagreement",
            "detail": f"spread_cents={row.get('spread_cents')}; models={row.get('models')}",
        })

    def sort_key(item: dict[str, Any]) -> tuple[int, float]:
        return (int(item.get("priority") or 9), float(item.get("time_sec") or 0))

    # Deduplicate same reason/time rounded to 0.1s.
    seen = set()
    deduped = []
    for row in sorted(review, key=sort_key):
        t = safe_float(row.get("time_sec"))
        key = (row.get("reason"), round(t or 0.0, 1))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped[:500]


def write_named_table_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[dict[str, Any]],
    chart_value_col: str | None = None,
    chart_title: str | None = None,
) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    set_basic_styles(ws)

    if chart_value_col and ws.max_row >= 3 and chart_value_col in headers:
        value_col = headers.index(chart_value_col) + 1
        chart = BarChart()
        chart.title = chart_title or title
        chart.y_axis.title = chart_value_col
        chart.x_axis.title = "Time sec"
        chart.height = 12
        chart.width = 24
        max_row = min(ws.max_row, 31)
        data = Reference(ws, min_col=value_col, min_row=1, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "J2")


def write_model_disagreement_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    write_named_table_sheet(
        wb,
        "Model Disagreement",
        ["time_sec", "support_count", "min_midi", "max_midi", "spread_cents", "suspected_octave_pair", "models"],
        rows,
        chart_value_col="spread_cents",
        chart_title="Raw model disagreements",
    )


def write_postprocess_delta_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    write_named_table_sheet(
        wb,
        "Postprocess Delta",
        ["time_sec", "target", "original_midi", "postprocessed_midi", "delta_cents", "change_type", "octave_like"],
        rows,
        chart_value_col="delta_cents",
        chart_title="Postprocess pitch changes",
    )


def write_final_compare_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    write_named_table_sheet(
        wb,
        "Final Compare",
        ["time_sec", "candidate_a", "candidate_b", "midi_a", "midi_b", "diff_cents", "relationship", "octave_like"],
        rows,
        chart_value_col="diff_cents",
        chart_title="Final RMVPE vs fusion differences",
    )


def write_fusion_fill_gap_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    write_named_table_sheet(
        wb,
        "Fusion Fill RMVPE Gap",
        [
            "time_sec",
            "fusion_midi",
            "fusion_postprocessed_midi",
            "hybrid_postprocessed_midi",
            "hybrid_decision",
            "support_count",
            "supporters",
            "torchcrepe_midi",
            "fcpe_midi",
            "pesto_midi",
            "risk",
        ],
        rows,
        chart_value_col=None,
    )


def write_review_segments_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    write_named_table_sheet(
        wb,
        "Review Segments",
        ["priority", "time_sec", "reason", "detail"],
        rows,
        chart_value_col=None,
    )


def write_diagnostics_sheet(wb: Workbook, json_files: list[Path]) -> None:
    ws = wb.create_sheet("Diagnostics")
    ws.append(["JSON file", "Key", "Value"])
    interesting_names = ("diagnostics", "fusion", "melody", "vocal_pitch", "postprocess", "cleaned")
    for path in sorted(json_files, key=lambda p: str(p)):
        if not any(name in path.name.lower() for name in interesting_names):
            continue
        data = load_json(path)
        if data is None:
            continue
        flat = flatten_dict(data, limit=250)
        for key, value in flat.items():
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            ws.append([str(path), key, value])
    set_basic_styles(ws)


def default_output_path(repo_root: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return repo_root / "tests" / "output" / f"fusion_report_{stamp}.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export latest yt2mp3 melody/fusion artifacts to Excel.")
    parser.add_argument("--repo-root", type=Path, default=None)
    parser.add_argument("--artifact-root", type=Path, action="append", default=[])
    parser.add_argument("--out", type=Path, default=None)
    parser.add_argument("--max-points", type=int, default=4000)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = find_repo_root(args.repo_root)
    log(f"repo root: {repo_root}")
    log(f"script version: {SCRIPT_VERSION}")

    search_roots = build_search_roots(repo_root, args.artifact_root)
    log("artifact search roots: " + ", ".join(str(p) for p in search_roots))

    group = discover_latest_group(search_roots)
    log(f"latest artifact group: {safe_rel(group.group_root, repo_root)}")
    if group.analysis_dir:
        log(f"analysis dir: {safe_rel(group.analysis_dir, repo_root)}")

    metadata_path, metadata = pick_metadata(group.json_files)
    if metadata_path:
        log(f"metadata source: {safe_rel(metadata_path, repo_root)}")
    else:
        log("YouTube metadata not found; Excel will still be generated.", "WARN")

    summaries = [summarize_csv(path) for path in sorted(group.csv_files, key=lambda p: (priority_for_csv(p), str(p)))]
    tracks, official_source = extract_official_tracks(group.csv_files)
    if not tracks:
        raise SystemExit(
            "有找到 CSV/JSON，但沒有辨識到可視化用的 official pitch 欄位。\n"
            "優先需要 comparison_postprocessed.csv 或 comparison.csv 內有 time/time_sec 與 fusion_f0_hz/rmvpe_f0_hz/torchcrepe_f0_hz/fcpe_f0_hz/pesto_f0_hz；若有 postprocess 欄位會額外加入 fusion_postprocessed/rmvpe_postprocessed。"
        )

    log(f"official comparison source: {official_source}")
    log("official pitch tracks: " + ", ".join(t.label for t in tracks))

    overlay_headers, overlay_rows = build_overlay_rows(tracks, max_points=max(200, args.max_points))

    # Legacy all-in-one views are kept for chart/reference only.
    # They are based on downsampled Pitch Overlay and should not drive decisions.
    disagreement_rows = build_disagreement_rows(overlay_headers, overlay_rows)
    octave_candidate_rows = build_octave_candidate_rows(overlay_headers, overlay_rows)

    # V4.2 split diagnostics use row-aligned source CSV directly.
    aligned_headers, aligned_rows, aligned_source = extract_row_aligned_official_rows(group.csv_files)
    if not aligned_rows:
        log(f"row-aligned diagnostics unavailable, falling back to Pitch Overlay: {aligned_source}", "WARN")
        aligned_headers, aligned_rows, aligned_source = overlay_headers, overlay_rows, "fallback: downsampled Pitch Overlay"

    model_headers, model_rows = filter_overlay_for_tracks(aligned_headers, aligned_rows, V4_MODEL_DISAGREEMENT_TRACKS)
    model_disagreement_rows = build_disagreement_rows(model_headers, model_rows)

    final_compare_rows = build_final_compare_rows(aligned_headers, aligned_rows)
    postprocess_delta_rows = build_postprocess_delta_rows(aligned_headers, aligned_rows)
    fusion_fill_gap_rows = build_fusion_fill_gap_rows(aligned_headers, aligned_rows)
    review_segment_rows = build_review_segment_rows(
        model_disagreement_rows=model_disagreement_rows,
        postprocess_delta_rows=postprocess_delta_rows,
        final_compare_rows=final_compare_rows,
        fusion_fill_gap_rows=fusion_fill_gap_rows,
    )
    model_summary_rows = build_model_summary_rows(group.json_files)

    log(f"V4.2 row-aligned diagnostics source: {aligned_source}")
    log(f"V4.2 row-aligned rows: {len(aligned_rows)}")
    log(f"V4.2 model disagreement rows: {len(model_disagreement_rows)}")
    log(f"V4.2 postprocess delta rows: {len(postprocess_delta_rows)}")
    log(f"V4.2 final compare rows: {len(final_compare_rows)}")
    log(f"V4.2 fusion fills RMVPE gap rows: {len(fusion_fill_gap_rows)}")
    log(f"V4.2 review segment rows: {len(review_segment_rows)}")

    wb = Workbook()
    write_summary(wb, group, repo_root, metadata_path, metadata, tracks, disagreement_rows, official_source)
    write_files_sheet(wb, group, repo_root)
    write_csv_index(wb, summaries)
    write_model_summary_sheet(wb, model_summary_rows)
    write_postprocess_summary_sheet(wb, group.json_files)

    # V4.2 split sheets for decision-making. These use row-aligned CSV rows.
    write_model_disagreement_sheet(wb, model_disagreement_rows)
    write_postprocess_delta_sheet(wb, postprocess_delta_rows)
    write_final_compare_sheet(wb, final_compare_rows)
    write_fusion_fill_gap_sheet(wb, fusion_fill_gap_rows)
    write_review_segments_sheet(wb, review_segment_rows)

    # Keep the original V3-style overview sheets as references.
    write_overlay_sheet(wb, overlay_headers, overlay_rows)
    write_disagreement_sheet(wb, disagreement_rows)
    write_octave_candidates_sheet(wb, octave_candidate_rows)
    write_diagnostics_sheet(wb, group.json_files)

    out_path = args.out or default_output_path(repo_root)
    out_path = out_path.expanduser()
    if not out_path.is_absolute():
        out_path = repo_root / out_path
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    log(f"Excel report written: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
